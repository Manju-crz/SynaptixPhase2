"""
Code Generator Utility - Dynamically generates pytest test files based on natural language queries
"""

import os
import re
import ast
import json
import logging
from typing import List, Dict, Optional, Any, Tuple
from datetime import datetime
from openpyxl import load_workbook

from loader.prompt_manager import save_prompts

logging.basicConfig(
    level=logging.INFO,
    format='\n%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)


class CodeGenerator:
    """
    Generates pytest test code dynamically based on natural language queries and API specifications.
    """

    def __init__(self, excel_path: str, base_url: str):
        """
        Initialize Code Generator.

        Args:
            excel_path: Path to Excel file with API specifications
            base_url: Base URL for API requests
        """
        self.excel_path = excel_path
        self.base_url = base_url
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.project_root = project_root
        self.rest_test_base = os.path.join(project_root, "rest_test")

        logger.info(f"🔧 Initializing Code Generator")
        logger.info(f"  Excel: {excel_path}")
        logger.info(f"  Base URL: {base_url}")

    @staticmethod
    def _derive_class_name(folder_name: str, filename: str) -> str:
        """
        Derive a valid Python class name from folder and file names.
        Combines folder and file names and strips any non-alphanumeric characters.

        Args:
            folder_name: Component/folder name
            filename: Test file name (without .py)

        Returns:
            A valid Python class identifier
        """
        raw = f"{folder_name}{filename}"
        name = re.sub(r'[^A-Za-z0-9]', '', raw)
        if not name or not name[0].isalpha():
            name = f"Test{name}"
        return name

    def _get_row_by_sl_no(self, sl_no: int) -> Optional[Dict[str, Any]]:
        """
        Get row data from Excel by Sl_No.

        Args:
            sl_no: Serial number to search for

        Returns:
            Dictionary with row data or None if not found
        """
        try:
            workbook = load_workbook(self.excel_path, data_only=True)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if row_dict.get('Sl_No') == sl_no:
                    logger.info(f"📋 Found row with Sl_No={sl_no}")
                    return row_dict

            logger.warning(f"⚠️  Row with Sl_No={sl_no} not found")
            return None

        except Exception as e:
            logger.error(f"❌ Error reading Excel: {str(e)}")
            return None

    def _parse_parameters(self, param_str: Optional[str]) -> Dict[str, Any]:
        """
        Parse parameter string from Excel format: name:#:REQ/NRQ:#:type:#:description
        Multiple parameters are separated by newlines.

        Args:
            param_str: Parameter string from Excel

        Returns:
            Dictionary of parameters with their details
        """
        if not param_str or str(param_str).strip() in ['', 'None', 'null']:
            return {}

        try:
            params = {}
            # Split by newline to get individual parameters
            param_lines = str(param_str).strip().split('\n')

            for line in param_lines:
                if not line.strip():
                    continue

                # Split by :#: delimiter
                parts = line.split(':#:')
                if len(parts) >= 2:
                    param_name = parts[0].strip()
                    is_required = parts[1].strip() == 'REQ' if len(parts) > 1 else False
                    param_type = parts[2].strip() if len(parts) > 2 else 'string'
                    param_desc = parts[3].strip() if len(parts) > 3 else ''

                    params[param_name] = {
                        'required': is_required,
                        'type': param_type,
                        'description': param_desc
                    }

            return params
        except Exception as e:
            logger.warning(f"⚠️  Could not parse parameters: {param_str}")
            return {}

    def _is_schema_definition(self, obj: Any) -> bool:
        """
        Check if object is an OpenAPI schema definition.

        Args:
            obj: Object to check

        Returns:
            True if it's a schema definition
        """
        if not isinstance(obj, dict):
            return False

        # Check for schema indicators
        schema_keys = {'type', 'properties', 'items', 'format', 'enum', 'required'}

        # If any value in the dict has 'type' field, it's likely a schema
        for value in obj.values():
            if isinstance(value, dict):
                if 'type' in value or 'properties' in value:
                    return True
                # Check for nested schema indicators
                if any(key in value for key in schema_keys):
                    return True

        # Check top level
        if any(key in obj for key in schema_keys):
            return True

        return False

    def _convert_schema_to_example(self, schema: Any) -> Any:
        """
        Convert OpenAPI schema definition to actual example value.

        Args:
            schema: Schema object (dict or primitive)

        Returns:
            Example value based on schema
        """
        if not isinstance(schema, dict):
            return schema

        # If it has 'example' field, use that
        if 'example' in schema:
            return schema['example']

        # If it has 'type' field, it's a schema definition
        if 'type' in schema:
            schema_type = schema['type']

            if schema_type == 'string':
                if 'enum' in schema:
                    return schema['enum'][0] if schema['enum'] else 'string'
                return schema.get('example', 'string_value')
            elif schema_type == 'integer':
                return schema.get('example', 0)
            elif schema_type == 'number':
                return schema.get('example', 0.0)
            elif schema_type == 'boolean':
                return schema.get('example', True)
            elif schema_type == 'array':
                items = schema.get('items', {})
                example_item = self._convert_schema_to_example(items)
                return [example_item]
            elif schema_type == 'object':
                # Handle object with properties
                if 'properties' in schema:
                    result = {}
                    for key, value in schema['properties'].items():
                        result[key] = self._convert_schema_to_example(value)
                    return result
                return {}

        # If it has 'properties', treat as object
        if 'properties' in schema:
            result = {}
            for key, value in schema['properties'].items():
                result[key] = self._convert_schema_to_example(value)
            return result

        # If it has 'items', treat as array
        if 'items' in schema:
            example_item = self._convert_schema_to_example(schema['items'])
            return [example_item]

        # Otherwise, recursively process all dict values
        result = {}
        for key, value in schema.items():
            if isinstance(value, dict):
                result[key] = self._convert_schema_to_example(value)
            else:
                result[key] = value
        return result

    def _parse_json_payload(self, json_str: Optional[str]) -> Optional[Dict]:
        """
        Parse JSON payload string from request_body_json column.
        Converts schema definitions to actual example values if needed.

        Args:
            json_str: JSON string from Excel (request_body_json column)

        Returns:
            Dictionary with actual example values or None
        """
        if not json_str or str(json_str).strip() in ['', 'None', 'null']:
            return None

        try:
            if isinstance(json_str, str):
                parsed = json.loads(json_str)

                # Check if it's a schema definition and convert to example
                if self._is_schema_definition(parsed):
                    logger.info("🔄 Converting schema definition to example value")
                    converted = self._convert_schema_to_example(parsed)
                    logger.info(f"✅ Converted to example: {json.dumps(converted, indent=2)[:200]}...")
                    return converted

                return parsed
            return None
        except Exception as e:
            logger.warning(f"⚠️  Could not parse JSON: {json_str}")
            return None

    def _format_json_payload(self, payload: Dict, indent_level: int = 3) -> str:
        """
        Format JSON payload as Python dict literal with proper indentation.
        Converts JSON booleans (true/false/null) to Python (True/False/None).

        Args:
            payload: Dictionary to format
            indent_level: Number of indentation levels (4 spaces each)

        Returns:
            Formatted string representation as valid Python code
        """
        import pprint

        # Use pprint to get Python representation (True/False/None instead of true/false/null)
        formatted = pprint.pformat(payload, indent=4, width=100, compact=False)

        lines = formatted.split('\n')

        # Add proper indentation to each line except the first
        indent = ' ' * (indent_level * 4)
        formatted_lines = []
        for i, line in enumerate(lines):
            if i == 0:
                formatted_lines.append(line)
            else:
                formatted_lines.append(indent + line)

        return '\n'.join(formatted_lines)

    def _sanitize_name(self, name: str) -> str:
        """
        Sanitize name to be a valid Python identifier.

        Args:
            name: Original name

        Returns:
            Sanitized name
        """
        # Replace spaces and special characters with underscores
        sanitized = ''.join(c if c.isalnum() or c == '_' else '_' for c in name.lower())
        # Remove consecutive underscores
        while '__' in sanitized:
            sanitized = sanitized.replace('__', '_')
        # Remove leading/trailing underscores
        sanitized = sanitized.strip('_')
        # Ensure it doesn't start with a number
        if sanitized and sanitized[0].isdigit():
            sanitized = 'test_' + sanitized
        return sanitized or 'test_api'

    def _parse_query_with_instructions(self, query: str) -> dict:
        """
        Parse query to extract main action and dependency instructions.

        Supports formats:
        1. "Create a new pet -> Retrieve the pet_id from the response"
        2. "Update pet -> Use the pet_id from previous response"
        3. "Delete a pet"

        Args:
            query: Natural language query with optional instructions

        Returns:
            Dictionary with:
            - 'action': Main query action
            - 'extract_fields': List of fields to extract from response
            - 'use_fields': List of fields to use from previous responses
            - 'has_dependency': Boolean
        """
        import re

        # Split by arrow to separate action from instructions
        if '->' in query:
            parts = [p.strip() for p in query.split('->')]
            action = parts[0]
            instructions = ' -> '.join(parts[1:])
        else:
            action = query.strip()
            instructions = ""

        result = {
            'action': action,
            'extract_fields': [],
            'use_fields': [],
            'has_dependency': False
        }

        if not instructions:
            return result

        # Parse extraction instructions: "Retrieve the pet_id from the response"
        extract_patterns = [
            r'retrieve\s+(?:the\s+)?(\w+)',
            r'extract\s+(?:the\s+)?(\w+)',
            r'save\s+(?:the\s+)?(\w+)',
            r'get\s+(?:the\s+)?(\w+)\s+from',
        ]

        for pattern in extract_patterns:
            matches = re.findall(pattern, instructions.lower())
            result['extract_fields'].extend(matches)

        # Parse usage instructions: "Use the pet_id from previous response"
        use_patterns = [
            r'use\s+(?:the\s+)?(\w+)',
            r'pass\s+(?:the\s+)?(\w+)',
            r'send\s+(?:the\s+)?(\w+)',
            r'with\s+(?:the\s+)?(\w+)\s+from',
        ]

        for pattern in use_patterns:
            matches = re.findall(pattern, instructions.lower())
            result['use_fields'].extend(matches)

        # Remove duplicates
        result['extract_fields'] = list(set(result['extract_fields']))
        result['use_fields'] = list(set(result['use_fields']))

        # Has dependency if using fields from previous responses
        result['has_dependency'] = len(result['use_fields']) > 0

        return result

    def _detect_dependency_in_query(self, query: str) -> dict:
        """
        Detect if query references data from previous steps.

        Args:
            query: Natural language query

        Returns:
            Dictionary with dependency info: {'has_dependency': bool, 'step_ref': int, 'field': str}
        """
        import re

        # First try to parse explicit instructions with arrows
        parsed = self._parse_query_with_instructions(query)
        if parsed['has_dependency']:
            return {
                'has_dependency': True,
                'step_ref': -1,  # Previous step
                'field': parsed['use_fields'][0] if parsed['use_fields'] else 'id',
                'extract_fields': parsed['extract_fields'],
                'use_fields': parsed['use_fields']
            }

        # Fallback to pattern-based detection
        patterns = [
            r'from step (\d+)',
            r'step (\d+)',
            r'from previous',
            r'created (\w+)',
            r'updated (\w+)',
            r'same (\w+)',
            r'above (\w+)',
        ]

        query_lower = query.lower()

        for pattern in patterns:
            match = re.search(pattern, query_lower)
            if match:
                if 'step' in pattern and match.group(1).isdigit():
                    return {'has_dependency': True, 'step_ref': int(match.group(1)), 'field': 'id', 'extract_fields': [], 'use_fields': []}
                elif 'previous' in query_lower:
                    return {'has_dependency': True, 'step_ref': -1, 'field': 'id', 'extract_fields': [], 'use_fields': []}
                else:
                    # Extract field name (pet, user, order, etc.)
                    field_match = match.group(1) if len(match.groups()) > 0 else 'id'
                    return {'has_dependency': True, 'step_ref': -1, 'field': field_match, 'extract_fields': [], 'use_fields': []}

        return {'has_dependency': False, 'step_ref': None, 'field': None, 'extract_fields': [], 'use_fields': []}

    def _generate_combined_test_method(self, queries: List[str], sl_nos: List[int],
                                        rows_data: List[Dict[str, Any]],
                                        method_index: int = 1) -> Tuple[str, str]:
        """
        Generate a single test method that executes multiple API calls sequentially.
        Automatically detects dependencies between steps based on query text.

        Args:
            queries: List of natural language queries
            sl_nos: List of serial numbers from Excel
            rows_data: List of row data from Excel
            method_index: 1-based index for the test method number (e.g., 1 -> test_01_)

        Returns:
            Tuple of (generated test method code as string, full method name)
        """
        # Parse first query to get clean action (without arrow instructions)
        first_parsed = self._parse_query_with_instructions(queries[0]) if queries else {'action': 'combined_api_test'}
        test_name = self._sanitize_name(first_parsed['action'])
        method_name = f"test_{method_index:02d}_{test_name}"

        # Build combined description using clean actions only
        combined_queries = '\n        '.join([
            f"{i+1}. {self._parse_query_with_instructions(q)['action']}"
            for i, q in enumerate(queries)
        ])

        # Build test code
        code = f'''    @allure.feature('API Testing')
    @allure.story('Combined API Test')
    @allure.severity(allure.severity_level.NORMAL)
    @allure.title('Combined Test: {len(queries)} API Operations')
    @allure.description("""
        This test executes {len(queries)} API operations sequentially:
        {combined_queries}
    """)
    def {method_name}(self, api_client):
        """
        Combined test executing {len(queries)} API operations
        """

'''

        # Generate code for each API call
        for i, (query, sl_no, row_data) in enumerate(zip(queries, sl_nos, rows_data), 1):
            method = str(row_data.get('Operation_Method', 'GET')).upper()
            path = row_data.get('Operation_Path', '/')
            operation_summary = row_data.get('Operation_Summary', query)

            # Detect if this query has dependencies on previous steps
            dependency_info = self._detect_dependency_in_query(query)

            # Parse all parameter types
            header_params = self._parse_parameters(row_data.get('header_parameters'))
            query_params = self._parse_parameters(row_data.get('query_parameters'))
            path_params = self._parse_parameters(row_data.get('path_parameters'))
            form_data_params = self._parse_parameters(row_data.get('form_data_parameters'))
            json_payload = self._parse_json_payload(row_data.get('request_body_json'))
            content_type = row_data.get('request_content_type', '')

            # Build endpoint with path params
            # Check if we can use response data from previous steps
            endpoint = path
            endpoint_uses_previous_data = False

            for param_name, param_info in path_params.items():
                # Check if query explicitly mentions dependency OR if it's a common ID pattern
                should_use_extracted = False
                extracted_var_name = None

                if dependency_info['has_dependency']:
                    # Query explicitly mentions using data from previous step
                    should_use_extracted = True

                    # Check if specific field is mentioned in use_fields
                    if dependency_info.get('use_fields'):
                        for use_field in dependency_info['use_fields']:
                            # Match field names (e.g., pet_id matches petId)
                            if use_field.lower().replace('_', '') == param_name.lower().replace('_', ''):
                                extracted_var_name = use_field
                                break

                    # If no specific match, use param_name
                    if not extracted_var_name:
                        extracted_var_name = param_name

                elif param_name.lower() in ['id', 'petid', 'userid', 'orderid', 'storeid'] and i > 1:
                    # Common ID pattern and not the first step
                    should_use_extracted = True
                    extracted_var_name = param_name

                if should_use_extracted:
                    # Use data from previous step's response
                    endpoint = endpoint.replace(f'{{{param_name}}}', f'{{extracted_{extracted_var_name}}}')
                    endpoint_uses_previous_data = True
                else:
                    placeholder = '1' if param_info['type'] in ['integer', 'int', 'number'] else 'value'
                    endpoint = endpoint.replace(f'{{{param_name}}}', placeholder)

            # Add step comment (use clean action only, without arrow instructions)
            clean_action = self._parse_query_with_instructions(query)['action']
            code += f'''        # Step {i}: {clean_action}
'''

            code += f'''        allure.dynamic.parameter("Step_{i}_Sl_No", {sl_no})
'''

            code += f'''        allure.dynamic.parameter("Step_{i}_Method", "{method}")
'''

            code += f'''        allure.dynamic.parameter("Step_{i}_Endpoint", "{path}")
'''

            code += f'''
'''

            # Add parameter variables with step prefix
            if path_params:
                code += f'''        # Step {i} - Path parameters
'''
                for param_name, param_info in path_params.items():
                    placeholder = '1' if param_info['type'] in ['integer', 'int', 'number'] else '"value"'
                    code += f'''        step{i}_{param_name} = {placeholder}  # {param_info['type']}
'''
                code += '\n'

            if header_params:
                code += f'''        # Step {i} - Headers
'''
                code += f'''        step{i}_headers = {{
'''
                for param_name, param_info in header_params.items():
                    req_marker = "# Required" if param_info['required'] else "# Optional"
                    code += f'''            "{param_name}": "value",  {req_marker}
'''
                code += f'''        }}

'''

            if query_params:
                code += f'''        # Step {i} - Query parameters
'''
                code += f'''        step{i}_params = {{
'''
                for param_name, param_info in query_params.items():
                    req_marker = "# Required" if param_info['required'] else "# Optional"
                    code += f'''            "{param_name}": "value",  {req_marker}
'''
                code += f'''        }}

'''

            if form_data_params:
                code += f'''        # Step {i} - Form data
'''
                code += f'''        step{i}_form_data = {{
'''
                for param_name, param_info in form_data_params.items():
                    req_marker = "# Required" if param_info['required'] else "# Optional"
                    code += f'''            "{param_name}": "value",  {req_marker}
'''
                code += f'''        }}

'''

            # Add endpoint variable
            code += f'''        step{i}_endpoint = "{endpoint}"

'''

            # Add API call with allure step
            code += f'''        with allure.step("Step {i}: {method} {path}"):
            logger.info(f"🚀 Step {i}: {{'{method}'}} {{step{i}_endpoint}}")
'''

            # Add payload if exists
            # Check content type to determine if it's JSON or form-urlencoded
            if json_payload:
                formatted_json = self._format_json_payload(json_payload)
                code += f'''            step{i}_payload = {formatted_json}
'''

            # Generate API call based on method
            if method == 'GET':
                code += f'''            response{i} = api_client.get(
'''
                code += f'''                endpoint=step{i}_endpoint'''
                if header_params:
                    code += f''',
                headers=step{i}_headers'''
                if query_params:
                    code += f''',
                params=step{i}_params'''
                code += '''
            )
'''

            elif method == 'POST':
                code += f'''            response{i} = api_client.post(
'''
                code += f'''                endpoint=step{i}_endpoint'''
                if header_params:
                    code += f''',
                headers=step{i}_headers'''
                if query_params:
                    code += f''',
                params=step{i}_params'''
                
                # Use data= for form-urlencoded, json_payload= for JSON
                if content_type == 'application/x-www-form-urlencoded':
                    if json_payload:
                        code += f''',
                data=step{i}_payload'''
                    elif form_data_params:
                        code += f''',
                data=step{i}_form_data'''
                else:
                    # Default to json_payload for application/json or other types
                    if json_payload:
                        code += f''',
                json_payload=step{i}_payload'''
                    elif form_data_params:
                        code += f''',
                data=step{i}_form_data'''
                
                code += '''
            )
'''

            elif method == 'PUT':
                code += f'''            response{i} = api_client.put(
'''
                code += f'''                endpoint=step{i}_endpoint'''
                if header_params:
                    code += f''',
                headers=step{i}_headers'''
                if query_params:
                    code += f''',
                params=step{i}_params'''
                
                # Use data= for form-urlencoded, json_payload= for JSON
                if content_type == 'application/x-www-form-urlencoded':
                    if json_payload:
                        code += f''',
                data=step{i}_payload'''
                    elif form_data_params:
                        code += f''',
                data=step{i}_form_data'''
                else:
                    # Default to json_payload for application/json or other types
                    if json_payload:
                        code += f''',
                json_payload=step{i}_payload'''
                    elif form_data_params:
                        code += f''',
                data=step{i}_form_data'''
                
                code += '''
            )
'''

            elif method == 'DELETE':
                code += f'''            response{i} = api_client.delete(
'''
                code += f'''                endpoint=step{i}_endpoint'''
                if header_params:
                    code += f''',
                headers=step{i}_headers'''
                if query_params:
                    code += f''',
                params=step{i}_params'''
                code += '''
            )
'''

            # Add logging for request payload
            if json_payload:
                code += f'''
            logger.info(f"📤 Step {i} Request Payload: {{json.dumps(step{i}_payload, indent=2)}}")
'''
            elif form_data_params:
                code += f'''
            logger.info(f"📤 Step {i} Form Data: {{step{i}_form_data}}")
'''
            
            # Add response verification and logging
            code += f'''
            logger.info(f"📥 Step {i} Response Status: {{response{i}.status_code}}")
            logger.info(f"📥 Step {i} Response Body: {{json.dumps(response{i}.json_data, indent=2) if response{i}.json_data else response{i}.text}}")
            
            allure.attach(
                json.dumps(response{i}.json_data, indent=2) if response{i}.json_data else response{i}.text,
                name="Step {i} Response",
                attachment_type=allure.attachment_type.JSON if response{i}.json_data else allure.attachment_type.TEXT
            )

            assert response{i}.is_success(), f"Step {i} failed with status {{response{i}.status_code}}"
            logger.info(f"✅ Step {i} passed - Status: {{response{i}.status_code}}")
'''

            # Extract fields from response based on query instructions
            parsed_query = self._parse_query_with_instructions(query)

            # Check if this step has explicit extraction instructions OR if subsequent steps need IDs
            needs_extraction = len(parsed_query['extract_fields']) > 0

            if not needs_extraction and method in ['POST', 'PUT'] and i < len(queries):
                # Check if any subsequent step needs an ID
                for future_idx in range(i, len(rows_data)):
                    future_query = queries[future_idx] if future_idx < len(queries) else ""
                    future_parsed = self._parse_query_with_instructions(future_query)

                    # Check explicit use_fields
                    if future_parsed['use_fields']:
                        needs_extraction = True
                        break

                    # Check path parameters
                    future_path_params = self._parse_parameters(rows_data[future_idx].get('path_parameters'))
                    for param_name in future_path_params.keys():
                        if param_name.lower() in ['id', 'petid', 'userid', 'orderid', 'storeid']:
                            needs_extraction = True
                            break
                    if needs_extraction:
                        break

            if needs_extraction:
                code += '''
            # Extract fields from response for use in subsequent steps
'''
                # Determine which fields to extract
                fields_to_extract = parsed_query['extract_fields'] if parsed_query['extract_fields'] else ['id']

                for field in fields_to_extract:
                    # NOTE: indentation here matches the enclosing `with allure.step(...)` body
                    # (12 spaces for if/else, 16 for their bodies) so the generated .py file
                    # is itself valid, importable Python.
                    code += f'''            if response{i}.json_data and '{field}' in response{i}.json_data:
                extracted_{field} = response{i}.json_data['{field}']
                logger.info(f"📌 Extracted {field}: {{extracted_{field}}}")
                allure.dynamic.parameter("Extracted_{field}", extracted_{field})
'''
                    # Alias lines are decided here, at generation time, based on the actual
                    # field name — NOT re-checked with a literal `field == ...` inside the
                    # generated code, since `field` is a generator-side variable and would
                    # be undefined (NameError) if referenced inside the emitted test method.
                    if field == 'id':
                        code += '''                # Aliases for common ID patterns
                extracted_petId = extracted_id
                extracted_userId = extracted_id
                extracted_orderId = extracted_id
'''
                    elif field == 'pet_id':
                        code += '''                extracted_petId = extracted_pet_id  # Alias
'''
                    elif field == 'user_id':
                        code += '''                extracted_userId = extracted_user_id  # Alias
'''

                    code += f'''            else:
                logger.warning("⚠️  '{field}' not found in response, using default value")
                extracted_{field} = 1 if '{field}' in ['id', 'pet_id', 'user_id', 'order_id'] else "default_value"
'''
                    if field == 'id':
                        code += '''                extracted_petId = extracted_id
                extracted_userId = extracted_id
                extracted_orderId = extracted_id
'''
                    code += '\n'

            code += f'''
'''

        return code, method_name

    def _generate_test_method(self, query: str, sl_no: int, row_data: Dict[str, Any], index: int) -> str:
        """
        Generate a single test method code.

        Args:
            query: Natural language query
            sl_no: Serial number from Excel
            row_data: Row data from Excel
            index: Test index number

        Returns:
            Generated test method code as string
        """
        method = str(row_data.get('Operation_Method', 'GET')).upper()
        path = row_data.get('Operation_Path', '/')
        operation_summary = row_data.get('Operation_Summary', query)

        # Parse all parameter types
        header_params = self._parse_parameters(row_data.get('header_parameters'))
        query_params = self._parse_parameters(row_data.get('query_parameters'))
        path_params = self._parse_parameters(row_data.get('path_parameters'))
        form_data_params = self._parse_parameters(row_data.get('form_data_parameters'))
        json_payload = self._parse_json_payload(row_data.get('request_body_json'))
        content_type = row_data.get('request_content_type', '')

        # Build endpoint with path params (use placeholder values for now)
        endpoint = path
        for param_name, param_info in path_params.items():
            # Use type-appropriate placeholder
            placeholder = '1' if param_info['type'] in ['integer', 'int', 'number'] else 'value'
            endpoint = endpoint.replace(f'{{{param_name}}}', placeholder)

        # Generate test method name
        test_name = self._sanitize_name(query)

        # Build parameter documentation
        param_docs = []
        if header_params:
            param_docs.append(f"    Headers: {list(header_params.keys())}")
        if query_params:
            param_docs.append(f"    Query Params: {list(query_params.keys())}")
        if path_params:
            param_docs.append(f"    Path Params: {list(path_params.keys())}")
        if form_data_params:
            param_docs.append(f"    Form Data: {list(form_data_params.keys())}")

        param_doc_str = '\n'.join(param_docs) if param_docs else '    No parameters'

        # Build test code
        code = f'''    @allure.feature('API Testing')
    @allure.story('{operation_summary}')
    @allure.severity(allure.severity_level.NORMAL)
    @allure.title('Test {index + 1}: {query}')
    @allure.description("""
        Query: {query}
        Sl_No: {sl_no}
        Method: {method}
        Endpoint: {path}
{param_doc_str}
    """)
    def test_{index + 1:02d}_{test_name}(self, api_client):
        """
        {query}

        Sl_No: {sl_no}
        Method: {method}
        Endpoint: {path}
        """
        allure.dynamic.parameter("Sl_No", {sl_no})
        allure.dynamic.parameter("Method", "{method}")
        allure.dynamic.parameter("Endpoint", "{path}")

        # TODO: Update parameter values as needed
'''

        # Add parameter variable definitions
        if path_params:
            code += f'''        # Path parameters
'''
            for param_name, param_info in path_params.items():
                placeholder = '1' if param_info['type'] in ['integer', 'int', 'number'] else '"value"'
                code += f'''        {param_name} = {placeholder}  # {param_info['type']} - {param_info['description']}
'''
            code += '\n'

        if header_params:
            code += f'''        # Header parameters
'''
            code += f'''        headers = {{
'''
            for param_name, param_info in header_params.items():
                req_marker = "# Required" if param_info['required'] else "# Optional"
                code += f'''            "{param_name}": "value",  {req_marker} - {param_info['description']}
'''
            code += f'''        }}

'''

        if query_params:
            code += f'''        # Query parameters
'''
            code += f'''        params = {{
'''
            for param_name, param_info in query_params.items():
                req_marker = "# Required" if param_info['required'] else "# Optional"
                placeholder = '1' if param_info['type'] in ['integer', 'int', 'number'] else '"value"'
                code += f'''            "{param_name}": {placeholder},  {req_marker} - {param_info['description']}
'''
            code += f'''        }}

'''

        if form_data_params:
            code += f'''        # Form data parameters
'''
            code += f'''        form_data = {{
'''
            for param_name, param_info in form_data_params.items():
                req_marker = "# Required" if param_info['required'] else "# Optional"
                code += f'''            "{param_name}": "value",  {req_marker} - {param_info['description']}
'''
            code += f'''        }}

'''

        # Build final endpoint with path params
        if path_params:
            code += f'''        # Build endpoint with path parameters
        endpoint = "{path}"
'''
            for param_name in path_params.keys():
                code += f'''        endpoint = endpoint.replace("{{{param_name}}}", str({param_name}))
'''
            code += '\n'
        else:
            code += f'''        endpoint = "{endpoint}"

'''

        # Add request execution based on method
        if method == 'GET':
            code += f'''        with allure.step("Execute GET request"):
            response = api_client.get(
                endpoint=endpoint'''
            if header_params:
                code += f''',
                headers=headers'''
            if query_params:
                code += f''',
                params=params'''
            code += '''
            )
'''

        elif method == 'POST':
            code += f'''        with allure.step("Execute POST request"):
'''
            if json_payload:
                formatted_json = self._format_json_payload(json_payload)
                code += f'''            payload = {formatted_json}

            response = api_client.post(
                endpoint=endpoint'''
                if header_params:
                    code += f''',
                headers=headers'''
                if query_params:
                    code += f''',
                params=params'''
                
                # Use data= for form-urlencoded, json_payload= for JSON
                if content_type == 'application/x-www-form-urlencoded':
                    code += f''',
                data=payload'''
                else:
                    code += f''',
                json_payload=payload'''
                
                code += '''
            )
'''
            else:
                code += f'''            response = api_client.post(
                endpoint=endpoint'''
                if header_params:
                    code += f''',
                headers=headers'''
                if query_params:
                    code += f''',
                params=params'''
                if form_data_params:
                    code += f''',
                data=form_data'''
                code += '''
            )
'''

        elif method == 'PUT':
            code += f'''        with allure.step("Execute PUT request"):
'''
            if json_payload:
                formatted_json = self._format_json_payload(json_payload)
                code += f'''            payload = {formatted_json}

            response = api_client.put(
                endpoint=endpoint'''
                if header_params:
                    code += f''',
                headers=headers'''
                if query_params:
                    code += f''',
                params=params'''
                
                # Use data= for form-urlencoded, json_payload= for JSON
                if content_type == 'application/x-www-form-urlencoded':
                    code += f''',
                data=payload'''
                else:
                    code += f''',
                json_payload=payload'''
                
                code += '''
            )
'''
            else:
                code += f'''            response = api_client.put(
                endpoint=endpoint'''
                if header_params:
                    code += f''',
                headers=headers'''
                if query_params:
                    code += f''',
                params=params'''
                if form_data_params:
                    code += f''',
                data=form_data'''
                code += '''
            )
'''

        elif method == 'DELETE':
            code += f'''        with allure.step("Execute DELETE request"):
            response = api_client.delete(
                endpoint=endpoint'''
            if header_params:
                code += f''',
                headers=headers'''
            if query_params:
                code += f''',
                params=params'''
            code += '''
            )
'''

        else:
            code += f'''        with allure.step("Execute {method} request"):
            # Unsupported method: {method}
            pytest.skip("Method {method} not implemented")
'''

        # Add assertions
        code += '''
        with allure.step("Verify response"):
            allure.attach(
                json.dumps(response.json_data, indent=2) if response.json_data else response.text,
                name="Response Body",
                attachment_type=allure.attachment_type.JSON if response.json_data else allure.attachment_type.TEXT
            )

            assert response.is_success(), f"Request failed with status {response.status_code}"
            logger.info(f"✅ Test passed - Status: {response.status_code}")
'''

        return code

    def _get_existing_test_method_names(self, file_path: str) -> List[str]:
        """
        Get all existing test method names from a Python test file using AST.

        Args:
            file_path: Path to the Python test file

        Returns:
            List of test method names (functions starting with 'test_')
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            tree = ast.parse(content)
            method_names = []
            for node in ast.walk(tree):
                if isinstance(node, ast.FunctionDef) and node.name.startswith('test_'):
                    method_names.append(node.name)
            return method_names
        except Exception as e:
            logger.warning(f"⚠️  Could not parse existing file for method names: {str(e)}")
            return []

    def _get_next_method_index(self, method_names: List[str]) -> int:
        """
        Determine the next test method index from existing method names.
        Parses the 'test_XX_' prefix to find the highest index.

        Args:
            method_names: List of existing test method names

        Returns:
            Next 1-based method index (e.g., if max is 2, returns 3)
        """
        max_index = 0
        pattern = re.compile(r'^test_(\d+)_')
        for name in method_names:
            match = pattern.match(name)
            if match:
                idx = int(match.group(1))
                if idx > max_index:
                    max_index = idx
        return max_index + 1

    def _insert_method_into_class(self, file_path: str, method_code: str) -> bool:
        """
        Insert a new test method into the first class found in the file.
        Uses AST to find the class's end line and inserts the method there.

        Args:
            file_path: Path to the Python test file
            method_code: The method code to insert (indented at class-body level, 4 spaces)

        Returns:
            True if insertion succeeded, False otherwise
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()

            tree = ast.parse(content)
            class_node = None
            for node in tree.body:
                if isinstance(node, ast.ClassDef):
                    class_node = node
                    break

            if not class_node:
                logger.warning("⚠️  Test class not found in existing file; cannot append")
                return False

            lines = content.split('\n')
            # end_lineno is 1-based; insert after the last line of the class body
            insert_at = class_node.end_lineno  # 0-based index for lines list

            # Ensure a blank line separator before the new method
            method_lines = method_code.split('\n')
            # Add a blank line before the new method for readability
            new_lines = lines[:insert_at] + [''] + method_lines + lines[insert_at:]
            new_content = '\n'.join(new_lines)

            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(new_content)

            logger.info(f"✅ Appended new test method into {class_node.name} class")
            return True

        except Exception as e:
            logger.error(f"❌ Error inserting method into class: {str(e)}")
            return False

    def generate_test_file(self, sl_nos: List[int], queries: List[str],
                            folder_name: str, filename: str,
                            original_query: str = None) -> Dict[str, Any]:
        """
        Generate a complete pytest test file based on queries and Sl_Nos.

        Args:
            sl_nos: List of serial numbers from Excel
            queries: List of natural language queries
            folder_name: Folder name to create inside rest_test/
            filename: Name of the test file (without .py extension)

        Returns:
            Dictionary with generation results
        """
        try:
            logger.info(f"\n{'='*80}")
            logger.info(f"🚀 GENERATING TEST FILE")
            logger.info(f"{'='*80}")
            logger.info(f"📁 Folder: {folder_name}")
            logger.info(f"📄 Filename: {filename}.py")
            logger.info(f"📊 Total Tests: {len(queries)}")
            logger.info(f"{'='*80}\n")

            # Create folder if not exists
            folder_path = os.path.join(self.rest_test_base, folder_name)
            os.makedirs(folder_path, exist_ok=True)
            logger.info(f"✅ Folder created/verified: {folder_path}")

            # Create __init__.py in folder if not exists
            init_file = os.path.join(folder_path, '__init__.py')
            if not os.path.exists(init_file):
                with open(init_file, 'w') as f:
                    f.write('"""\nGenerated test package\n"""\n')
                logger.info(f"✅ Created __init__.py")

            # Check if the target file already exists (append mode)
            file_path = os.path.join(folder_path, f"{filename}.py")
            file_exists = os.path.exists(file_path)

            # Get all row data
            rows_data = []
            generated_tests = []
            for i, (query, sl_no) in enumerate(zip(queries, sl_nos)):
                logger.info(f"  Loading data {i+1}/{len(queries)}: {query} (Sl_No: {sl_no})")

                row_data = self._get_row_by_sl_no(sl_no)
                if not row_data:
                    logger.warning(f"⚠️  Skipping query '{query}' - Sl_No {sl_no} not found")
                    continue

                rows_data.append(row_data)
                generated_tests.append({
                    'query': query,
                    'sl_no': sl_no,
                    'method': row_data.get('Operation_Method'),
                    'endpoint': row_data.get('Operation_Path')
                })

            # Generate single combined test method for all queries
            logger.info(f"Generating combined test method for {len(queries)} queries")

            generated_method_names = []

            # Derive the class name from the folder and file names
            class_name = self._derive_class_name(folder_name, filename)
            logger.info(f"🏷️  Derived test class name: {class_name}")

            if file_exists:
                # === APPEND MODE: add new method to existing class ===
                logger.info(f"📎 Existing file detected: {file_path} — appending new test method")

                existing_methods = self._get_existing_test_method_names(file_path)
                method_index = self._get_next_method_index(existing_methods)
                logger.info(f"   Existing methods: {len(existing_methods)} | Next method index: {method_index:02d}")

                if rows_data:
                    test_code, method_name = self._generate_combined_test_method(
                        queries, sl_nos, rows_data, method_index=method_index
                    )
                    inserted = self._insert_method_into_class(file_path, test_code)
                    if not inserted:
                        # Fallback: could not insert into class, treat as failure
                        return {
                            'success': False,
                            'error': 'Could not append method to existing TestGeneratedAPIs class',
                            'message': 'Failed to append method to existing test class'
                        }
                    generated_method_names.append(method_name)
                    logger.info(f"   ✅ Appended method: {method_name}")
            else:
                # === CREATE MODE: generate fresh file with header, fixture, and class ===
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                file_content = f'''"""
Generated Test File
Generated on: {timestamp}
Base URL: {self.base_url}
Excel Source: {os.path.basename(self.excel_path)}
"""

import json
import pytest
import allure
import logging
from rest_util.rest_api_client import RestApiClient

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@pytest.fixture(scope="class")
def api_client():
    """Fixture to provide REST API client"""
    client = RestApiClient(base_url="{self.base_url}")
    yield client
    client.close()

@allure.suite('Generated API Tests')
class {class_name}:
    """
    Auto-generated test class based on natural language queries
    """

'''

                if rows_data:
                    test_code, method_name = self._generate_combined_test_method(
                        queries, sl_nos, rows_data, method_index=1
                    )
                    file_content += test_code + '\n'
                    generated_method_names.append(method_name)

                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(file_content)

            logger.info(f"\n{'='*80}")
            logger.info(f"✅ TEST FILE GENERATED SUCCESSFULLY")
            logger.info(f"{'='*80}")
            logger.info(f"📁 Location: {file_path}")
            logger.info(f"📄 Tests Generated: {len(generated_tests)}")
            logger.info(f"{'='*80}\n")

            # Save the original prompts for each generated method
            prompt_text = original_query if original_query else '; '.join(queries)
            prompts = {name: prompt_text for name in generated_method_names}
            save_prompts(self.project_root, folder_name, filename, class_name, prompts)

            return {
                'success': True,
                'file_path': file_path,
                'folder_path': folder_path,
                'class_name': class_name,
                'tests_generated': len(generated_tests),
                'tests': generated_tests,
                'generated_method_names': generated_method_names,
                'message': f'Successfully generated {len(generated_tests)} tests'
            }

        except Exception as e:
            logger.error(f"❌ Error generating test file: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': str(e),
                'message': f'Failed to generate test file: {str(e)}'
            }


def generate_tests_from_queries(excel_path: str, base_url: str, sl_nos: List[int],
                                  queries: List[str], folder_name: str, filename: str) -> Dict[str, Any]:
    """
    Convenience function to generate test file from queries.

    Args:
        excel_path: Path to Excel file with API specifications
        base_url: Base URL for API requests
        sl_nos: List of serial numbers from Excel
        queries: List of natural language queries
        folder_name: Folder name to create inside rest_test/
        filename: Name of the test file (without .py extension)

    Returns:
        Dictionary with generation results
    """
    generator = CodeGenerator(excel_path=excel_path, base_url=base_url)
    return generator.generate_test_file(sl_nos, queries, folder_name, filename)