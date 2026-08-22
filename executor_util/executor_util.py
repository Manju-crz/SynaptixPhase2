"""
API Executor Utility - Integrates Semantic Search with REST API Execution
"""

import json
import logging
from typing import Dict, List, Optional, Any, Union
from openpyxl import load_workbook

from nlp.semantic_search_util import SemanticSearchEngine
from rest_util.rest_api_client import RestApiClient
from rest_util.config import BASE_URLS, TIMEOUTS

logging.basicConfig(
    level=logging.INFO,
    format='\n%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)


class ApiExecutor:
    """
    API Executor that uses semantic search to find relevant API endpoints
    and executes them using REST API client.
    """

    def __init__(self, excel_path: str, base_url: str,
                 search_columns: Optional[List[str]] = None,
                 default_timeout: int = None):
        """
        Initialize API Executor.

        Args:
            excel_path: Path to Excel file with API specifications
            base_url: Base URL for API requests
            search_columns: Columns to use for semantic search
            default_timeout: Default timeout for API requests
        """
        self.excel_path = excel_path
        self.base_url = base_url
        self.default_timeout = default_timeout or TIMEOUTS['DEFAULT']

        logger.info(f"🔧 Initializing API Executor")
        logger.info(f"   Excel: {excel_path}")
        logger.info(f"   Base URL: {base_url}")

        if search_columns is None:
            search_columns = ['Component', 'Component_SmallDescription', 'Operation_Method',
                              'Operation_Path', 'Operation_Summary', 'Operation_SecondarySummary']

        self.search_engine = SemanticSearchEngine(excel_path, search_columns=search_columns)
        self.api_client = RestApiClient(base_url=base_url, timeout=self.default_timeout)

        logger.info(f"✅ API Executor initialized successfully")

    def _get_row_by_sl_no(self, sl_no: int) -> Optional[Dict[str, Any]]:
        """
        Get row data from Excel by Sl_No.

        Args:
            sl_no: Serial number to search for

        Returns:
            Dictionary with row data or None if not found
        """
        try:
            workbook = load_workbook(self.excel_path, data_only=True)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if row_dict.get('Sl_No') == sl_no:
                    logger.info(f"📄 Found row with Sl_No={sl_no}")
                    return row_dict

            logger.warning(f"⚠️ Row with Sl_No={sl_no} not found")
            return None

        except Exception as e:
            logger.error(f"❌ Error reading Excel: {str(e)}")
            return None

    def _parse_parameters(self, param_str: Optional[str]) -> Dict[str, Any]:
        """
        Parse parameter string (JSON or comma-separated) into dictionary.

        Args:
            param_str: Parameter string from Excel

        Returns:
            Dictionary of parameters
        """
        if not param_str or str(param_str).strip() in ['', 'None', 'null']:
            return {}

        try:
            if isinstance(param_str, str) and param_str.strip().startswith('{'):
                return json.loads(param_str)
            else:
                params = {}
                for item in str(param_str).split(','):
                    if ':' in item:
                        key, value = item.split(':', 1)
                        params[key.strip()] = value.strip()
                return params
        except Exception as e:
            logger.warning(f"⚠️ Could not parse parameters: {param_str} - {str(e)}")
            return {}

    def _convert_schema_to_example(self, schema: Any) -> Any:
        """
        Convert OpenAPI schema definition to actual example value.

        Args:
            schema: Schema object (dict or primitive)

        Returns:
            Example value based on schema
        """
        if not isinstance(schema, dict):
            return schema

        # If it has 'example' field, use that
        if 'example' in schema:
            return schema['example']

        # If it has 'type' field, it's a schema definition
        if 'type' in schema:
            schema_type = schema['type']

            if schema_type == 'string':
                return schema.get('example', 'string')
            elif schema_type == 'integer':
                return schema.get('example', 0)
            elif schema_type == 'number':
                return schema.get('example', 0.0)
            elif schema_type == 'boolean':
                return schema.get('example', True)
            elif schema_type == 'array':
                items = schema.get('items', {})
                example_item = self._convert_schema_to_example(items)
                return [example_item]
            elif schema_type == 'object':
                # Handle object with properties
                if 'properties' in schema:
                    result = {}
                    for key, value in schema['properties'].items():
                        result[key] = self._convert_schema_to_example(value)
                    return result
                return {}

        # If it has 'properties', treat as object
        if 'properties' in schema:
            result = {}
            for key, value in schema['properties'].items():
                result[key] = self._convert_schema_to_example(value)
            return result

        # If it has 'items', treat as array
        if 'items' in schema:
            example_item = self._convert_schema_to_example(schema['items'])
            return [example_item]

        # Otherwise, recursively process all dict values
        result = {}
        for key, value in schema.items():
            if isinstance(value, dict):
                result[key] = self._convert_schema_to_example(value)
            else:
                result[key] = value
        return result

    def _is_schema_definition(self, obj: Any) -> bool:
        """
        Check if object is an OpenAPI schema definition.

        Args:
            obj: Object to check

        Returns:
            True if it's a schema definition
        """
        if not isinstance(obj, dict):
            return False

        # Check for schema indicators
        schema_keys = {'type', 'properties', 'items', 'format', 'enum', 'required'}

        # If any value in the dict has 'type' field, it's likely a schema
        for value in obj.values():
            if isinstance(value, dict):
                if 'type' in value or 'properties' in value:
                    return True
                # Check for nested schema indicators
                if any(key in value for key in schema_keys):
                    return True

        # Check top level
        if any(key in obj for key in schema_keys):
            return True

        return False

    def _parse_json_payload(self, json_str: Optional[str]) -> Optional[Dict]:
        """
        Parse JSON payload string and convert schema to example if needed.

        Args:
            json_str: JSON string from Excel

        Returns:
            Dictionary or None
        """
        if not json_str or str(json_str).strip() in ['', 'None', 'null']:
            return None

        try:
            if isinstance(json_str, str):
                parsed = json.loads(json_str)
                # Check if it's a schema definition
                if self._is_schema_definition(parsed):
                    logger.info("   🔄 Converting schema definition to example value")
                    converted = self._convert_schema_to_example(parsed)
                    logger.info(f"   ✅ Converted payload: {json.dumps(converted, indent=2)}")
                    return converted
                return parsed
            return None
        except Exception as e:
            logger.warning(f"⚠️ Could not parse JSON: {json_str} - {str(e)}")
            return None

    def _build_endpoint(self, path: str, path_params: Dict[str, Any]) -> str:
        """
        Build endpoint URL by replacing path parameters.

        Args:
            path: API path with placeholders (e.g., /pet/{petId})
            path_params: Dictionary of path parameters

        Returns:
            Formatted endpoint path
        """
        endpoint = path
        for key, value in path_params.items():
            endpoint = endpoint.replace(f'{{{key}}}', str(value))
        return endpoint

    def execute_api_call(self, query: str, timeout: Optional[int] = None) -> Dict[str, Any]:
        """
        Execute API call based on natural language query.

        Args:
            query: Natural language query
            timeout: Optional timeout override

        Returns:
            Dictionary with execution results
        """
        logger.info(f"🔍 Processing query: '{query}'")

        sl_no = self.search_engine.get_best_match_sl_no(query)

        if sl_no is None:
            logger.error(f"❌ No matching API found for query: '{query}'")
            return {
                'success': False,
                'query': query,
                'error': 'No matching API found',
                'sl_no': None
            }

        logger.info(f"✅ Found matching API: Sl_No={sl_no}")

        row_data = self._get_row_by_sl_no(sl_no)

        if not row_data:
            return {
                'success': False,
                'query': query,
                'error': f'Row data not found for Sl_No={sl_no}',
                'sl_no': sl_no
            }

        method = str(row_data.get('Operation_Method', 'GET')).upper()
        path = row_data.get('Operation_Path', '/')

        header_params = self._parse_parameters(row_data.get('header_parameters'))
        query_params = self._parse_parameters(row_data.get('query_parameters'))
        path_params = self._parse_parameters(row_data.get('path_parameters'))
        form_params = self._parse_parameters(row_data.get('form_data_parameters'))
        json_payload = self._parse_json_payload(row_data.get('request_body_json'))

        endpoint = self._build_endpoint(path, path_params)

        logger.info(f"\n{'='*80}")
        logger.info(f"🚀 EXECUTING API CALL")
        logger.info(f"{'='*80}")
        logger.info(f"◆ Method: {method}")
        logger.info(f"◆ Endpoint: {endpoint}")
        logger.info(f"◆ Full URL: {self.base_url}{endpoint}")

        if header_params:
            logger.info(f"\n📑 Request Headers:")
            for key, value in header_params.items():
                logger.info(f"   {key}: {value}")

        if query_params:
            logger.info(f"\n🔍 Query Parameters:")
            for key, value in query_params.items():
                logger.info(f"   {key}: {value}")

        if path_params:
            logger.info(f"\n🖼️ Path Parameters:")
            for key, value in path_params.items():
                logger.info(f"   {key}: {value}")

        if json_payload:
            logger.info(f"\n📦 REQUEST BODY (JSON Payload):")
            logger.info(json.dumps(json_payload, indent=2))

        if form_params:
            logger.info(f"\n📝 Form Data Parameters:")
            for key, value in form_params.items():
                logger.info(f"   {key}: {value}")

        if not json_payload and not form_params:
            logger.info(f"\n📦 REQUEST BODY: None")

        logger.info(f"{'='*80}\n")

        try:
            request_timeout = timeout or self.default_timeout

            if method == 'GET':
                response = self.api_client.get(
                    endpoint=endpoint,
                    headers=header_params if header_params else None,
                    params=query_params if query_params else None
                )
            elif method == 'POST':
                response = self.api_client.post(
                    endpoint=endpoint,
                    headers=header_params if header_params else None,
                    params=query_params if query_params else None,
                    json_payload=json_payload,
                    data=form_params if form_params else None
                )
            elif method == 'PUT':
                response = self.api_client.put(
                    endpoint=endpoint,
                    headers=header_params if header_params else None,
                    params=query_params if query_params else None,
                    json_payload=json_payload,
                    data=form_params if form_params else None
                )
            elif method == 'PATCH':
                response = self.api_client.patch(
                    endpoint=endpoint,
                    headers=header_params if header_params else None,
                    params=query_params if query_params else None,
                    json_payload=json_payload,
                    data=form_params if form_params else None
                )
            elif method == 'DELETE':
                response = self.api_client.delete(
                    endpoint=endpoint,
                    headers=header_params if header_params else None,
                    params=query_params if query_params else None,
                    json_payload=json_payload
                )
            else:
                logger.error(f"❌ Unsupported HTTP method: {method}")
                return {
                    'success': False,
                    'query': query,
                    'error': f'Unsupported HTTP method: {method}',
                    'sl_no': sl_no,
                    'method': method
                }

            result = {
                'success': response.is_success(),
                'query': query,
                'sl_no': sl_no,
                'method': method,
                'endpoint': endpoint,
                'status_code': response.status_code,
                'response_time': response.elapsed_time,
                'response_data': response.json_data,
                'response_text': response.text if not response.json_data else None,
                'headers': dict(response.headers),
                'request_details': {
                    'header_params': header_params,
                    'query_params': query_params,
                    'path_params': path_params,
                    'form_params': form_params,
                    'json_payload': json_payload
                }
            }

            logger.info(f"✅ API call completed: Status={response.status_code}, Time={response.elapsed_time:.2f}s")

            return result

        except Exception as e:
            logger.error(f"❌ API call failed: {str(e)}")
            return {
                'success': False,
                'query': query,
                'error': str(e),
                'sl_no': sl_no,
                'method': method,
                'endpoint': endpoint
            }

    def execute_multiple_queries(self, queries: List[str],
                                 timeout: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Execute multiple API calls based on list of natural language queries.

        Args:
            queries: List of natural language queries
            timeout: Optional timeout override

        Returns:
            List of execution results
        """
        logger.info(f"🚀 Executing {len(queries)} queries")

        results = []
        for i, query in enumerate(queries, 1):
            logger.info(f"\n{'='*80}")
            logger.info(f"Query {i}/{len(queries)}")
            logger.info(f"{'='*80}")

            result = self.execute_api_call(query, timeout)
            results.append(result)

        logger.info(f"\n{'='*80}")
        logger.info(f"✅ Completed {len(queries)} queries")
        logger.info(f"   Successful: {sum(1 for r in results if r['success'])}")
        logger.info(f"   Failed: {sum(1 for r in results if not r['success'])}")
        logger.info(f"{'='*80}")

        return results

    def close(self):
        """Close API client session"""
        self.api_client.close()
        logger.info("🔒 API Executor closed")


def execute_single_query(excel_path: str, base_url: str, query: str,
                         search_columns: Optional[List[str]] = None,
                         timeout: Optional[int] = None) -> Dict[str, Any]:
        """
        Convenience function to execute a single query.

        Args:
            excel_path: Path to Excel file
            base_url: Base URL for API
            query: Natural language query
            search_columns: Columns for semantic search
            timeout: Request timeout

        Returns:
            Execution result dictionary
        """
        executor = ApiExecutor(excel_path, base_url, search_columns, timeout)
        result = executor.execute_api_call(query, timeout)
        executor.close()
        return result


def execute_multiple_queries_batch(excel_path: str, base_url: str, queries: List[str],
                                   search_columns: Optional[List[str]] = None,
                                   timeout: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Convenience function to execute multiple queries.

        Args:
            excel_path: Path to Excel file
            base_url: Base URL for API
            queries: List of natural language queries
            search_columns: Columns for semantic search
            timeout: Request timeout

        Returns:
            List of execution results
        """
        executor = ApiExecutor(excel_path, base_url, search_columns, timeout)
        results = executor.execute_multiple_queries(queries, timeout)
        executor.close()
        return results