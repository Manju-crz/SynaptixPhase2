"""
Parameter Extractor Utility - Extract API parameters from Excel by Sl_No
Extracts header, query, path, form data parameters, operation path, example JSON, and response JSON
"""

import os
import json
import logging
from typing import Dict, Optional, Any
from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)


class ParameterExtractor:
    """
    Extracts API parameters and details from Excel file based on SL_No.
    """

    def __init__(self, excel_path: str):
        """
        Initialize Parameter Extractor.

        Args:
            excel_path: Path to Excel file with API specifications
        """
        self.excel_path = excel_path
        logger.info(f"🔷 Initializing Parameter Extractor")
        logger.info(f"   Excel: {excel_path}")

    def _get_row_by_sl_no(self, sl_no: int) -> Optional[Dict[str, Any]]:
        """
        Get row data from Excel by SL_No.

        Args:
            sl_no: Serial number to search for

        Returns:
            Dictionary with row data or None if not found
        """
        try:
            workbook = load_workbook(self.excel_path, data_only=True)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if row_dict.get('Sl_No') == sl_no:
                    logger.info(f"🟢 Found row with Sl_No={sl_no}")
                    return row_dict

            logger.warning(f"⚠️ Row with Sl_No={sl_no} not found")
            return None

        except Exception as e:
            logger.error(f"❌ Error reading Excel: {str(e)}")
            return None

    def _parse_json_string(self, json_str: Optional[str]) -> Optional[Dict]:
        """
        Parse JSON string into dictionary.

        Args:
            json_str: JSON string from Excel

        Returns:
            Dictionary or None
        """
        if not json_str or str(json_str).strip() in ['', 'None', 'null']:
            return None

        try:
            if isinstance(json_str, str):
                return json.loads(json_str)
            return None
        except Exception as e:
            logger.warning(f"⚠️ Could not parse JSON: {json_str}")
            return None

    def _parse_parameters(self, param_str: Optional[str]) -> Dict[str, Any]:
        """
        Parse parameter string (JSON or comma-separated) into dictionary.

        Args:
            param_str: Parameter string from Excel

        Returns:
            Dictionary of parameters
        """
        if not param_str or str(param_str).strip() in ['', 'None', 'null']:
            return {}

        try:
            if isinstance(param_str, str) and param_str.strip().startswith('{'):
                return json.loads(param_str)
            return {}
        except Exception as e:
            logger.warning(f"⚠️ Could not parse parameters: {param_str}")
            return {}

    def extract_parameters(self, sl_no: int) -> Dict[str, Any]:
        """
        Extract all API parameters and details for a given SL_No.

        Args:
            sl_no: Serial number from Excel

        Returns:
            Dictionary containing all extracted parameters and details
        """
        logger.info(f"\n{'='*80}")
        logger.info(f"🔽 EXTRACTING PARAMETERS FOR SL_No: {sl_no}")
        logger.info(f"{'='*80}")

        # Get row data
        row_data = self._get_row_by_sl_no(sl_no)

        if not row_data:
            return {
                'success': False,
                'error': f'SL_No {sl_no} not found in Excel',
                'sl_no': sl_no
            }

        # Extract all parameters
        extracted_data = {
            'success': True,
            'sl_no': sl_no,
            'operation_method': row_data.get('Operation_Method'),
            'operation_path': row_data.get('Operation_Path'),
            'operation_summary': row_data.get('Operation_Summary'),
            'operation_description': row_data.get('Operation_SecondanySummary'),
            'header_parameters': self._parse_parameters(row_data.get('header_parameters')),
            'query_parameters': self._parse_parameters(row_data.get('query_parameters')),
            'path_parameters': self._parse_parameters(row_data.get('path_parameters')),
            'form_data_parameters': self._parse_parameters(row_data.get('form_data_parameters')),
            'request_body_json': self._parse_json_string(row_data.get('request_body_json')),
            'response_json': self._parse_json_string(row_data.get('response_model_json')),
            'response_code': row_data.get('response_code'),
            'tags': row_data.get('tags')
        }

        return extracted_data

    def print_parameters(self, sl_no: int):
        """
        Extract and print all API parameters for a given SL_No.

        Args:
            sl_no: Serial number from Excel
        """
        data = self.extract_parameters(sl_no)

        if not data['success']:
            logger.error(f"❌ {data['error']}")
            return

        # Print extracted data
        logger.info(f"\n{'='*80}")
        logger.info(f"📋 EXTRACTED PARAMETERS")
        logger.info(f"{'='*80}")

        logger.info(f"\n🔹 Basic Information:")
        logger.info(f"   SL_No: {data['sl_no']}")
        logger.info(f"   Method: {data['operation_method']}")
        logger.info(f"   Path: {data['operation_path']}")
        logger.info(f"   Summary: {data['operation_summary']}")
        if data['operation_description']:
            logger.info(f"   Description: {data['operation_description']}")
        if data['tags']:
            logger.info(f"   Tags: {data['tags']}")

        # Header Parameters
        if data['header_parameters']:
            logger.info(f"\n📊 Header Parameters:")
            logger.info(json.dumps(data['header_parameters'], indent=4))
        else:
            logger.info(f"\n📊 Header Parameters: None")

        # Query Parameters
        if data['query_parameters']:
            logger.info(f"\n🔷 Query Parameters:")
            logger.info(json.dumps(data['query_parameters'], indent=4))
        else:
            logger.info(f"\n🔷 Query Parameters: None")

        # Path Parameters
        if data['path_parameters']:
            logger.info(f"\n🟦 Path Parameters:")
            logger.info(json.dumps(data['path_parameters'], indent=4))
        else:
            logger.info(f"\n🟦 Path Parameters: None")

        # Form Data Parameters
        if data['form_data_parameters']:
            logger.info(f"\n📝 Form Data Parameters:")
            logger.info(json.dumps(data['form_data_parameters'], indent=4))
        else:
            logger.info(f"\n📝 Form Data Parameters: None")

        # Request Body JSON
        if data['request_body_json']:
            logger.info(f"\n🔸 Request Body JSON:")
            logger.info(json.dumps(data['request_body_json'], indent=4))
        else:
            logger.info(f"\n🔸 Request Body JSON: None")

        # Response JSON
        if data['response_json']:
            logger.info(f"\n🔵 Response JSON:")
            logger.info(json.dumps(data['response_json'], indent=4))
        else:
            logger.info(f"\n🔵 Response JSON: None")

        # Response Code
        if data['response_code']:
            logger.info(f"\n✅ Response Code: {data['response_code']}")

        logger.info(f"\n{'='*80}\n")


def extract_and_print_parameters(excel_path: str, sl_no: int):
    """
    Convenience function to extract and print parameters.

    Args:
        excel_path: Path to Excel file
        sl_no: Serial number to extract
    """
    extractor = ParameterExtractor(excel_path)
    extractor.print_parameters(sl_no)


def extract_parameters_only(excel_path: str, sl_no: int) -> Dict[str, Any]:
    """
    Convenience function to extract parameters without printing.

    Args:
        excel_path: Path to Excel file
        sl_no: Serial number to extract

    Returns:
        Dictionary with extracted parameters
    """
    extractor = ParameterExtractor(excel_path)
    return extractor.extract_parameters(sl_no)
