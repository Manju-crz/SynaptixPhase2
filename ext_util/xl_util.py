"""
Excel Utilities - Reusable Excel file operations using openpyxl
"""

import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from .filesys_util import create_folder_if_not_exists

logging.basicConfig(
    level=logging.INFO,
    format='\n%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)


def create_excel_file(folder_path: str, filename: str, sheet_name: str = "Sheet1") -> str:
    """
    Create a new Excel (.xlsx) file.

    Args:
        folder_path: Path to the folder where file will be created
        filename: Name of the file (without extension)
        sheet_name: Name of the default sheet (default: "Sheet1")

    Returns:
        str: Full path to the created Excel file
    """
    # Ensure folder exists
    create_folder_if_not_exists(folder_path)

    # Build full file path
    full_filename = f"{filename}.xlsx"
    file_path = os.path.join(folder_path, full_filename)

    # Create workbook and set sheet name
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # Save the workbook
    workbook.save(file_path)
    logger.info(f"✅ Excel file created: {file_path}")
    return file_path


def write_to_excel(file_path: str, sheet_name: str, data: list) -> None:
    """
    Write data to an existing Excel file.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to write to
        data: List of lists containing data to write
    """
    # Load the existing workbook
    workbook = load_workbook(file_path)

    # Get or create the sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Write data to the sheet
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # Save the workbook
    workbook.save(file_path)
    logger.info(f"✅ Data written to Excel file: {file_path} (Sheet: {sheet_name})")


def create_excel_with_data(folder_path: str, filename_prefix: str, sheet_name: str, columns: list, data: list) -> str:
    """
    Create a new Excel file with timestamp and write column headers and data.

    Args:
        folder_path: Path to the folder where file will be created
        filename_prefix: Prefix for the filename (timestamp will be appended)
        sheet_name: Name of the sheet
        columns: List of column headers
        data: List of data rows (each row is a list)

    Returns:
        str: Full path to the created Excel file
    """
    # Generate timestamp
    timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
    filename = f"{filename_prefix}_{timestamp}"

    # Create Excel file
    file_path = create_excel_file(folder_path, filename, sheet_name)

    # Write column headers and data
    write_to_excel(file_path, sheet_name, [columns] + data)

    logger.info(f"✅ Excel file created with {len(data)} data rows: {file_path}")
    return file_path