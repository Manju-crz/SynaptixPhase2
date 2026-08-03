"""
Semantic Search Utility - Find similar rows in Excel using sentence transformers
"""

import os
import logging
from typing import List, Dict, Tuple, Optional
import numpy as np
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer, util

logging.basicConfig(
    level=logging.INFO,
    format='\n%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)


class SemanticSearchEngine:
    """
    Semantic search engine using all-mpnet-base-v2 model for finding similar rows in Excel.
    """

    def __init__(self, excel_path: str, model_name: str = 'all-mpnet-base-v2', search_columns: Optional[List[str]] = None):
        """
        Initialize the semantic search engine.

        Args:
            excel_path: Path to the Excel file
            model_name: Name of the sentence transformer model (default: 'all-mpnet-base-v2')
            search_columns: List of column names to use for semantic search. If None, uses all columns except Sl_No
        """
        self.excel_path = excel_path
        self.model_name = model_name
        self.model = None
        self.data = []
        self.embeddings = None
        self.columns = []
        self.search_columns = search_columns

        logger.info(f"🔧 Initializing Semantic Search Engine with model: {model_name}")
        self._load_model()
        self._load_excel_data()
        self._generate_embeddings()

    def _load_model(self):
        """Load the sentence transformer model."""
        try:
            logger.info(f"📦 Loading model: {self.model_name}")
            self.model = SentenceTransformer(self.model_name)
            logger.info(f"✅ Model loaded successfully")
        except Exception as e:
            logger.error(f"❌ Failed to load model: {str(e)}")
            raise

    def _load_excel_data(self):
        """Load data from Excel file."""
        try:
            logger.info(f"📁 Loading Excel file: {self.excel_path}")

            if not os.path.exists(self.excel_path):
                raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

            workbook = load_workbook(self.excel_path, read_only=True)
            sheet = workbook.active

            rows = list(sheet.iter_rows(values_only=True))

            if len(rows) == 0:
                raise ValueError("Excel file is empty")

            self.columns = list(rows[0])

            for row in rows[1:]:
                row_dict = {}
                for idx, col_name in enumerate(self.columns):
                    row_dict[col_name] = row[idx] if idx < len(row) else None
                self.data.append(row_dict)

            workbook.close()

            logger.info(f"✅ Loaded {len(self.data)} rows with {len(self.columns)} columns")
            logger.info(f"📋 Columns: {', '.join(self.columns)}")

        except Exception as e:
            logger.error(f"❌ Failed to load Excel data: {str(e)}")
            raise

    def _generate_embeddings(self, text_columns: Optional[List[str]] = None):
        """
        Generate embeddings for all rows.

        Args:
            text_columns: List of column names to use for generating embeddings.
                          If None, uses search_columns or all columns except Sl_No.
        """
        try:
            logger.info("🔵 Generating embeddings for all rows...")

            if text_columns is None:
                if self.search_columns is not None:
                    text_columns = self.search_columns
                else:
                    text_columns = [col for col in self.columns if col not in ['Sl_No', 'Sl No', 'sl_no', 'SL_NO']]
            else:
                invalid_cols = [col for col in text_columns if col not in self.columns]
                if invalid_cols:
                    raise ValueError(f"Invalid columns: {invalid_cols}")

            texts = []
            for row in self.data:
                row_text_parts = []
                for col in text_columns:
                    value = row.get(col)
                    if value is not None and str(value).strip():
                        row_text_parts.append(f"{col}: {str(value)}")

                combined_text = " | ".join(row_text_parts)
                texts.append(combined_text)

            self.embeddings = self.model.encode(texts, convert_to_tensor=True, show_progress_bar=True)

            logger.info(f"✅ Generated embeddings for {len(texts)} rows")

        except Exception as e:
            logger.error(f"❌ Failed to generate embeddings: {str(e)}")
            raise

    def search(self, query: str, top_k: int = 1) -> List[Dict]:
        """
        Search for the most similar rows based on natural language query.

        Args:
            query: Natural language search query
            top_k: Number of top results to return (default: 1)

        Returns:
            List of dictionaries containing:
                - sl_no: Serial number from the Excel (if 'Sl No' column exists)
                - row_index: 0-based index in the data list
                - excel_row: Actual Excel row number (1-based, accounting for header)
                - score: Similarity score (0-1)
                - data: Row data as dictionary
        """
        try:
            if self.embeddings is None:
                raise ValueError("Embeddings not generated. Call _generate_embeddings first.")

            logger.info(f"🔍 Searching for: '{query}'")

            query_embedding = self.model.encode(query, convert_to_tensor=True)

            cos_scores = util.cos_sim(query_embedding, self.embeddings)[0]

            top_results = np.argsort(-cos_scores.cpu().numpy())[:top_k]

            results = []
            for idx in top_results:
                row_data = self.data[idx]
                score = float(cos_scores[idx])

                sl_no = None
                if 'Sl_No' in row_data:
                    sl_no = row_data['Sl_No']
                elif 'Sl No' in row_data:
                    sl_no = row_data['Sl No']
                elif 'sl_no' in row_data:
                    sl_no = row_data['sl_no']
                elif 'SL_NO' in row_data:
                    sl_no = row_data['SL_NO']

                result = {
                    'sl_no': sl_no,
                    'row_index': int(idx),
                    'excel_row': int(idx) + 2,
                    'score': score,
                    'data': row_data
                }
                results.append(result)

                logger.info(f"📋 Result #{len(results)}: Sl No={sl_no}, Score={score:.4f}, Excel Row={int(idx) + 2}")

            return results

        except Exception as e:
            logger.error(f"❌ Search failed: {str(e)}")
            raise

    def get_best_match_sl_no(self, query: str) -> Optional[int]:
        """
        Get the Sl No of the best matching row.

        Args:
            query: Natural language search query

        Returns:
            Sl No of the best matching row, or None if not found
        """
        results = self.search(query, top_k=1)
        if results and len(results) > 0:
            best_match = results[0]
            logger.info(f"✅ Best match Sl No: {best_match['sl_no']} (Score: {best_match['score']:.4f})")
            return best_match['sl_no']
        return None

    def regenerate_embeddings(self, text_columns: List[str]):
        """
        Regenerate embeddings using specific columns.

        Args:
            text_columns: List of column names to use for generating embeddings
        """
        logger.info(f"🔵 Regenerating embeddings with columns: {', '.join(text_columns)}")
        self._generate_embeddings(text_columns)


def find_similar_row(
        excel_path: str, query: str, model_name: str = 'all-mpnet-base-v2', search_columns: Optional[List[str]] = None) -> Optional[int]:
    """
    Convenience function to find the most similar row and return its Sl No.

    Args:
        excel_path: Path to the Excel file
        query: Natural language search query
        model_name: Name of the sentence transformer model (default: 'all-mpnet-base-v2')
        search_columns: List of column names to use for semantic search. If None, uses all columns except Sl_No

    Returns:
        Sl No of the most similar row, or None if not found
    """
    engine = SemanticSearchEngine(excel_path, model_name, search_columns)
    return engine.get_best_match_sl_no(query)


def find_top_k_similar_rows(excel_path: str, query: str, top_k: int = 5,
                             model_name: str = 'all-mpnet-base-v2', search_columns: Optional[List[str]] = None) -> List[Dict]:
    """
    Convenience function to find top K similar rows.

    Args:
        excel_path: Path to the Excel file
        query: Natural language search query
        top_k: Number of top results to return
        model_name: Name of the sentence transformer model (default: 'all-mpnet-base-v2')
        search_columns: List of column names to use for semantic search. If None, uses all columns except Sl_No

    Returns:
        List of result dictionaries
    """
    engine = SemanticSearchEngine(excel_path, model_name, search_columns)
    return engine.search(query, top_k)